import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

wb = openpyxl.load_workbook('phl5_compliance.xlsx', data_only=True, read_only=True)
print('Sheets:', wb.sheetnames, flush=True)

ws = wb.active
print('Active sheet:', ws.title, flush=True)
print('Max rows:', ws.max_row, flush=True)

header_found = False
data_count = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if not header_found:
        print(f'Row {i}:', row[:5], flush=True)
        if row[0] == 'Associate':
            header_found = True
            print('HEADER FOUND at row', i, flush=True)
    else:
        if row[0] and data_count < 3:
            print('DATA:', row[:5], flush=True)
            data_count += 1
        if i > 20:
            break

wb.close()
print('DONE', flush=True)