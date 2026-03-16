import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('phl5_compliance_copy.xlsx', data_only=True)
print('All sheets:', wb.sheetnames, flush=True)

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f'\n=== {sname} === rows:{ws.max_row} cols:{ws.max_column}', flush=True)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 15:
            print(row, flush=True)
        else:
            break